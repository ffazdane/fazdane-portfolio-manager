"""
Tax Center — Excel Report Generator
=====================================
Builds an in-memory Excel workbook with multiple tabs:

  1. Executive Summary
  2. Ticker Summary
  3. Broker Summary
  4. Section 1256 Summary
  5. Detailed Transactions
  6. Extraction Warnings

Returns bytes suitable for st.download_button.
No database writes; purely in-memory.
"""

from __future__ import annotations
import io
from collections import defaultdict
from typing import Any

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

from src.tax.normalizer import normalise_ticker, is_section_1256, split_6040

# ── Style palette ─────────────────────────────────────────────────────────────
CLR_DARK_BG   = "0E1117"
CLR_HEADER    = "00D4AA"
CLR_HEADER_TXT= "0E1117"
CLR_ROW_EVEN  = "1A1F2E"
CLR_ROW_ODD   = "141924"
CLR_WARNING   = "FF4B4B"
CLR_POSITIVE  = "00D4AA"
CLR_NEGATIVE  = "FF4B4B"
CLR_NEUTRAL   = "AAAAAA"
CLR_SEC1256   = "FFA421"

_THIN = Side(border_style="thin", color="2A2F3E")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

MONEY_FMT = '#,##0.00;[Red]-#,##0.00'
DATE_FMT  = 'YYYY-MM-DD'


def _hdr_font(bold: bool = True) -> Font:
    return Font(name="Calibri", bold=bold, color=CLR_HEADER_TXT, size=10)


def _body_font(color: str = "E0E0E0", bold: bool = False) -> Font:
    return Font(name="Calibri", color=color, bold=bold, size=10)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _write_header_row(ws, row_idx: int, headers: list[str]):
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=row_idx, column=ci, value=h)
        cell.font      = _hdr_font()
        cell.fill      = _fill(CLR_HEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _BORDER


def _write_data_row(ws, row_idx: int, values: list[Any], money_cols: set[int] = None,
                    highlight_color: str = None):
    bg = CLR_ROW_EVEN if row_idx % 2 == 0 else CLR_ROW_ODD
    if highlight_color:
        bg = highlight_color
    for ci, val in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=ci, value=val)
        cell.fill   = _fill(bg)
        cell.border = _BORDER
        cell.alignment = Alignment(vertical="center")
        if money_cols and ci in money_cols:
            cell.number_format = MONEY_FMT
            if isinstance(val, (int, float)):
                color = CLR_POSITIVE if val >= 0 else CLR_NEGATIVE
                cell.font = _body_font(color=color)
            else:
                cell.font = _body_font()
        else:
            cell.font = _body_font()


def _autofit(ws, min_width: int = 10, max_width: int = 45):
    for col_cells in ws.columns:
        max_len = max_width
        for cell in col_cells:
            try:
                clen = len(str(cell.value or ""))
                if clen > max_len:
                    max_len = clen
            except Exception:
                pass
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(
            max(min_width, max_len + 2), max_width
        )


# ── Aggregation helpers ───────────────────────────────────────────────────────

def _aggregate_by_ticker(txns: list[dict]) -> list[dict]:
    """Roll up transactions to one row per (normalized_symbol, broker, account)."""
    buckets: dict[tuple, dict] = defaultdict(lambda: {
        "qty_bought": 0.0, "total_buy": 0.0, "qty_sold": 0.0, "total_proceeds": 0.0,
        "realized_pl": 0.0, "st_pl": 0.0, "lt_pl": 0.0,
        "sec1256_pl": 0.0, "lt_60": 0.0, "st_40": 0.0,
        "brokers": set(), "accounts": set(),
    })

    for t in txns:
        sym    = t.get("normalized_symbol") or t.get("original_symbol") or "UNKNOWN"
        broker = t.get("broker", "")
        acct   = t.get("account", "")
        key    = (sym, broker, acct)
        b      = buckets[key]

        qty   = t.get("quantity") or 0.0
        proc  = t.get("proceeds", 0.0)
        cost  = t.get("cost_basis", 0.0)
        gl    = t.get("gain_loss", 0.0)
        term  = t.get("term", "UNKNOWN")
        s1256 = t.get("section_1256", False)

        b["qty_sold"]       += qty
        b["total_proceeds"] += proc
        b["total_buy"]      += cost
        b["realized_pl"]    += gl
        b["brokers"].add(broker)
        b["accounts"].add(acct)

        if term == "SHORT":
            b["st_pl"] += gl
        elif term == "LONG":
            b["lt_pl"] += gl

        if s1256:
            b["sec1256_pl"] += gl
            lt60, st40 = split_6040(gl)
            b["lt_60"] += lt60
            b["st_40"] += st40

    rows = []
    for (sym, broker, acct), b in sorted(buckets.items()):
        avg_sell = b["total_proceeds"] / b["qty_sold"] if b["qty_sold"] else 0.0
        avg_buy  = b["total_buy"]      / b["qty_sold"] if b["qty_sold"] else 0.0
        tax_cat  = "Section 1256" if b["sec1256_pl"] != 0 else (
            "Long-term" if b["lt_pl"] != 0 else (
                "Short-term" if b["st_pl"] != 0 else "Mixed/Unknown"
            )
        )
        rows.append({
            "Ticker":           sym,
            "Broker":           broker,
            "Account":          acct,
            "Qty Sold":         round(b["qty_sold"], 4),
            "Avg Sell Price":   round(avg_sell, 4),
            "Total Proceeds":   round(b["total_proceeds"], 2),
            "Avg Buy Price":    round(avg_buy, 4),
            "Total Cost Basis": round(b["total_buy"], 2),
            "Realized P/L":     round(b["realized_pl"], 2),
            "Short-Term P/L":   round(b["st_pl"], 2),
            "Long-Term P/L":    round(b["lt_pl"], 2),
            "Sec 1256 P/L":     round(b["sec1256_pl"], 2),
            "60% LT (Sec1256)": round(b["lt_60"], 2),
            "40% ST (Sec1256)": round(b["st_40"], 2),
            "Tax Category":     tax_cat,
            "Notes":            "",
        })
    return rows


def _aggregate_by_broker(txns: list[dict]) -> list[dict]:
    buckets: dict[tuple, dict] = defaultdict(lambda: {
        "proceeds": 0.0, "cost": 0.0, "gl": 0.0, "st": 0.0, "lt": 0.0, "s1256": 0.0
    })
    for t in txns:
        key = (t.get("broker", ""), t.get("account", ""), t.get("tax_year", ""))
        b = buckets[key]
        b["proceeds"] += t.get("proceeds", 0.0)
        b["cost"]     += t.get("cost_basis", 0.0)
        b["gl"]       += t.get("gain_loss", 0.0)
        if t.get("term") == "SHORT": b["st"] += t.get("gain_loss", 0.0)
        if t.get("term") == "LONG":  b["lt"] += t.get("gain_loss", 0.0)
        if t.get("section_1256"):    b["s1256"] += t.get("gain_loss", 0.0)
    return [
        {
            "Broker":          broker,
            "Account":         acct,
            "Tax Year":        year,
            "Total Proceeds":  round(b["proceeds"], 2),
            "Total Cost Basis":round(b["cost"],     2),
            "Realized P/L":    round(b["gl"],        2),
            "Short-Term P/L":  round(b["st"],        2),
            "Long-Term P/L":   round(b["lt"],        2),
            "Sec 1256 P/L":    round(b["s1256"],     2),
        }
        for (broker, acct, year), b in sorted(buckets.items())
    ]


def _sec1256_summary(txns: list[dict]) -> list[dict]:
    buckets: dict[tuple, dict] = defaultdict(lambda: {"gl": 0.0, "lt60": 0.0, "st40": 0.0})
    for t in txns:
        if not t.get("section_1256"):
            continue
        sym  = t.get("normalized_symbol") or t.get("original_symbol") or "UNKNOWN"
        key  = (sym, t.get("broker", ""), t.get("account", ""))
        gl   = t.get("gain_loss", 0.0)
        lt60, st40 = split_6040(gl)
        buckets[key]["gl"]   += gl
        buckets[key]["lt60"] += lt60
        buckets[key]["st40"] += st40
    return [
        {
            "Normalized Ticker":        sym,
            "Broker":                   broker,
            "Account":                  acct,
            "Total Sec 1256 P/L":       round(b["gl"],   2),
            "60% Long-Term Allocation": round(b["lt60"], 2),
            "40% Short-Term Allocation":round(b["st40"], 2),
            "Notes": "Verify against Form 6781 and tax advisor",
        }
        for (sym, broker, acct), b in sorted(buckets.items())
    ]


# ── Main entry ────────────────────────────────────────────────────────────────

def build_excel_report(
    transactions: list[dict],
    warnings: list[str],
    tax_year: str = "Unknown",
) -> bytes:
    """Build and return the Excel workbook as raw bytes."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    # ── Computed aggregates ───────────────────────────────────────────────────
    ticker_rows  = _aggregate_by_ticker(transactions)
    broker_rows  = _aggregate_by_broker(transactions)
    sec1256_rows = _sec1256_summary(transactions)

    total_proceeds = sum(t.get("proceeds",   0.0) for t in transactions)
    total_cost     = sum(t.get("cost_basis", 0.0) for t in transactions)
    total_gl       = sum(t.get("gain_loss",  0.0) for t in transactions)
    total_st       = sum(t.get("gain_loss",  0.0) for t in transactions if t.get("term") == "SHORT")
    total_lt       = sum(t.get("gain_loss",  0.0) for t in transactions if t.get("term") == "LONG")
    total_s1256    = sum(t.get("gain_loss",  0.0) for t in transactions if t.get("section_1256"))
    s1256_lt60, s1256_st40 = split_6040(total_s1256)

    brokers  = sorted({t.get("broker",  "") for t in transactions})
    accounts = sorted({t.get("account", "") for t in transactions})
    years    = sorted({t.get("tax_year","") for t in transactions})

    # ──────────────────────────────────────────────────────────────────────────
    # Tab 1: Executive Summary
    # ──────────────────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Executive Summary")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 25

    summary_rows = [
        ("Tax Year",                     ", ".join(years) or tax_year),
        ("Brokers",                       ", ".join(brokers)),
        ("Accounts",                      ", ".join(accounts)),
        ("",                              ""),
        ("Total Proceeds",               round(total_proceeds, 2)),
        ("Total Cost Basis",             round(total_cost,     2)),
        ("Total Realized Gain / Loss",   round(total_gl,       2)),
        ("",                              ""),
        ("Short-Term Gain / Loss",       round(total_st,       2)),
        ("Long-Term Gain / Loss",        round(total_lt,       2)),
        ("",                              ""),
        ("Section 1256 Gain / Loss",     round(total_s1256,    2)),
        ("  60% Long-Term (Sec 1256)",   round(s1256_lt60,     2)),
        ("  40% Short-Term (Sec 1256)",  round(s1256_st40,     2)),
        ("",                              ""),
        ("Total Transactions Extracted", len(transactions)),
        ("Extraction Warnings",          len(warnings)),
    ]

    title_cell = ws.cell(row=1, column=1, value="FazDane Analytics — Tax Center Report")
    title_cell.font = Font(name="Calibri", bold=True, size=14, color=CLR_HEADER)
    title_cell.fill = _fill(CLR_DARK_BG)
    ws.merge_cells("A1:B1")

    for ri, (label, value) in enumerate(summary_rows, 3):
        lc = ws.cell(row=ri, column=1, value=label)
        vc = ws.cell(row=ri, column=2, value=value)
        lc.fill = _fill(CLR_ROW_EVEN if ri % 2 == 0 else CLR_ROW_ODD)
        vc.fill = _fill(CLR_ROW_EVEN if ri % 2 == 0 else CLR_ROW_ODD)
        lc.font = _body_font(color="AAAAAA")
        lc.border = _BORDER
        vc.border = _BORDER
        if isinstance(value, float):
            vc.number_format = MONEY_FMT
            color = CLR_POSITIVE if value >= 0 else CLR_NEGATIVE
            vc.font = _body_font(color=color, bold=True)
        else:
            vc.font = _body_font(bold=True)

    ws.row_dimensions[1].height = 28

    # ──────────────────────────────────────────────────────────────────────────
    # Tab 2: Ticker Summary
    # ──────────────────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Ticker Summary")
    ws2.sheet_view.showGridLines = False
    if ticker_rows:
        headers = list(ticker_rows[0].keys())
        _write_header_row(ws2, 1, headers)
        money_cols = {i+1 for i, h in enumerate(headers)
                      if any(k in h for k in ["Price","Proceeds","Cost","P/L","60%","40%","Sec"])}
        for ri, row in enumerate(ticker_rows, 2):
            vals = list(row.values())
            hi_color = None
            if row.get("Tax Category") == "Section 1256":
                hi_color = "1A1510"  # very subtle orange tint
            _write_data_row(ws2, ri, vals, money_cols=money_cols, highlight_color=hi_color)
        _autofit(ws2)

    # ──────────────────────────────────────────────────────────────────────────
    # Tab 3: Broker Summary
    # ──────────────────────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Broker Summary")
    ws3.sheet_view.showGridLines = False
    if broker_rows:
        headers = list(broker_rows[0].keys())
        _write_header_row(ws3, 1, headers)
        money_cols = {i+1 for i, h in enumerate(headers)
                      if any(k in h for k in ["Proceeds","Cost","P/L"])}
        for ri, row in enumerate(broker_rows, 2):
            _write_data_row(ws3, ri, list(row.values()), money_cols=money_cols)
        _autofit(ws3)

    # ──────────────────────────────────────────────────────────────────────────
    # Tab 4: Section 1256 Summary
    # ──────────────────────────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Section 1256 Summary")
    ws4.sheet_view.showGridLines = False
    disclaimer = ws4.cell(row=1, column=1,
        value="⚠  Section 1256 classification is preliminary. "
              "Verify against Form 6781 and consult your tax advisor.")
    disclaimer.font = Font(name="Calibri", bold=True, color=CLR_WARNING, size=10)
    disclaimer.fill = _fill("1F0A0A")
    ws4.merge_cells(f"A1:{get_column_letter(7)}1")
    ws4.row_dimensions[1].height = 20

    if sec1256_rows:
        headers = list(sec1256_rows[0].keys())
        _write_header_row(ws4, 2, headers)
        money_cols = {i+1 for i, h in enumerate(headers) if "P/L" in h or "Allocation" in h}
        for ri, row in enumerate(sec1256_rows, 3):
            _write_data_row(ws4, ri, list(row.values()), money_cols=money_cols)
        _autofit(ws4)
    else:
        ws4.cell(row=3, column=1, value="No Section 1256 transactions detected.").font = _body_font(color="888888")

    # ──────────────────────────────────────────────────────────────────────────
    # Tab 5: Detailed Transactions
    # ──────────────────────────────────────────────────────────────────────────
    ws5 = wb.create_sheet("Detailed Transactions")
    ws5.sheet_view.showGridLines = False
    detail_headers = [
        "Broker","Account","Tax Year","Source File",
        "Original Symbol","Normalized Symbol","CUSIP","Description",
        "Date Acquired","Date Sold","Quantity",
        "Proceeds","Cost Basis","Wash Sale Adj","Gain / Loss",
        "Term","Section 1256","60% LT","40% ST",
        "Confidence","Review Note",
    ]
    _write_header_row(ws5, 1, detail_headers)
    money_cols5 = {12, 13, 14, 15, 18, 19}
    for ri, t in enumerate(transactions, 2):
        gl = t.get("gain_loss", 0.0)
        lt60, st40 = split_6040(gl) if t.get("section_1256") else (0.0, 0.0)
        conf = t.get("extraction_confidence", "")
        hi = "1F0A0A" if conf == "LOW" else None
        vals = [
            t.get("broker",""),        t.get("account",""),
            t.get("tax_year",""),      t.get("source_file",""),
            t.get("original_symbol",""), t.get("normalized_symbol",""),
            t.get("cusip",""),         t.get("description","")[:100],
            t.get("date_acquired",""), t.get("date_sold",""),
            t.get("quantity",""),
            t.get("proceeds",0.0),     t.get("cost_basis",0.0),
            t.get("wash_sale_adj",0.0),gl,
            t.get("term",""),          "Yes" if t.get("section_1256") else "No",
            lt60, st40,
            conf,                      t.get("review_note",""),
        ]
        _write_data_row(ws5, ri, vals, money_cols=money_cols5, highlight_color=hi)
    ws5.freeze_panes = "A2"
    _autofit(ws5, max_width=50)

    # ──────────────────────────────────────────────────────────────────────────
    # Tab 6: Extraction Warnings
    # ──────────────────────────────────────────────────────────────────────────
    ws6 = wb.create_sheet("Extraction Warnings")
    ws6.sheet_view.showGridLines = False
    _write_header_row(ws6, 1, ["#", "Warning"])
    ws6.column_dimensions["A"].width = 6
    ws6.column_dimensions["B"].width = 100
    if warnings:
        for ri, w in enumerate(warnings, 2):
            ws6.cell(row=ri, column=1, value=ri-1).font = _body_font(color="888888")
            c = ws6.cell(row=ri, column=2, value=w)
            c.font = _body_font(color=CLR_WARNING)
            c.fill = _fill(CLR_ROW_EVEN if ri%2==0 else CLR_ROW_ODD)
            c.border = _BORDER
            c.alignment = Alignment(wrap_text=True)
    else:
        ws6.cell(row=2, column=2, value="No extraction warnings 🎉").font = _body_font(color=CLR_POSITIVE)

    # ── Freeze header rows ────────────────────────────────────────────────────
    for sheet in [ws2, ws3, ws4, ws5]:
        sheet.freeze_panes = "A2"

    # ── Serialise ─────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
